#!/usr/bin/env python
#  midnam2xlsx.py
#
#  Copyright 2020 Christopher Arndt <chris@chrisarndt.de>
#
"""Convert MIDNAM instrument description into a patch list excel spread sheet."""

import argparse
import logging
import sys
from collections import namedtuple
import xml.etree.ElementTree as etree

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .styles import colhead, numcell, txtcell


log = logging.getLogger("midnam2xlsx")

COLUMNS_ALL = {
    "A": ("Bank", 8, colhead, numcell),
    "B": ("MSB", 7, colhead, numcell),
    "C": ("LSB", 7, colhead, numcell),
    "D": ("PC", 7, colhead, numcell),
    "E": ("Category", 12, colhead, numcell),
    "F": ("Name", 20, colhead, txtcell),
}

COLUMNS_BANK = {
    "A": ("PC", 6, colhead, numcell),
    "B": ("Category", 12, colhead, numcell),
    "C": ("Name", 20, colhead, txtcell),
}

CATEGORIES = {
    "Br": "Brass",
    "Bs": "Bass",
    "Cp": "Chromatic Perc",
    "Dr": "Drum / Perc",
    "Et": "Ethnic",
    "Gt": "Guitar",
    "Kb": "Keyboard",
    "Ld": "Syn Lead",
    "Me": "Musical FX",
    "Or": "Organ",
    "Pd": "Pad / Choir",
    "Pn": "Piano",
    "Sc": "Syn Comp",
    "Se": "Sound FX",
    "St": "Strings",
    "WW": "Woodwind",
}


PatchBank = namedtuple("PatchBank", ("name", "msb", "lsb", "patchlist"))
Patch = namedtuple("Patch", ("name", "category", "program"))


def sanitize(s):
    for repl in r"\/*?:[]":
        s = s.replace(repl, " ")
    return s


def parse_patchnamelist(patchnamelist):
    patchlist = []

    for pnum, patch in enumerate(patchnamelist.findall("./Patch"), start=1):
        pnumber = patch.get("Number", pnum)
        pname = patch.get("Name", "<unnamed patch #{}".format(patch.get("Number", pnum)))

        # If name is the same as the patch number, skip this patch
        try:
            if int(pnumber) == int(pname):
                continue
        except (TypeError, ValueError):
            pass

        try:
            program = int(patch.get("ProgramChange"))
        except:
            program = ""

        try:
            category, pname = pname.split(":", 1)
        except:
            category = ""
        else:
            category = CATEGORIES.get(category, category)

        patchlist.append(Patch(pname, category, program))

    return patchlist


def parse_midnam(fileobj):
    tree = etree.parse(fileobj)
    root = tree.getroot()

    if root.tag != "MIDINameDocument":
        raise ValueError("Unrecognized document type '%s'." % root.tag)

    patchlists = {}
    for plnum, patchnamelist in enumerate(root.findall("./MasterDeviceNames/PatchNameList"), start=1):
        plname = patchnamelist.get("Name", "<unnamed patchlist #{}>".format(plnum))
        log.debug("Patch Name List: %s", plname)
        patchlists[plname] = parse_patchnamelist(patchnamelist)

    namesets = {}
    for nsnum, nameset in enumerate(root.findall("./MasterDeviceNames/ChannelNameSet"), start=1):
        nsname = nameset.get("Name", "<unnamed set #{}>".format(nsnum))
        log.debug("Channel Name Set: %s", nsname)

        patchbanks = []

        for pbnum, patchbank in enumerate(nameset.findall("./PatchBank")):
            pbname = patchbank.get("Name", "<unnamed bank #{}>".format(pbnum + 1))
            log.debug("Patch Bank: %s", pbname)

            msb = patchbank.find("./MIDICommands/ControlChange[@Control='0']")

            try:
                msb = int(msb.get("Value"))
            except:
                msb = ""

            lsb = patchbank.find("./MIDICommands/ControlChange[@Control='32']")

            try:
                lsb = int(lsb.get("Value"))
            except:
                lsb = ""

            patchlist_ref = patchbank.find("./UsesPatchNameList")

            if patchlist_ref is not None:
                plname = patchlist_ref.get("Name")
                try:
                    patchbanks.append(PatchBank(pbname, msb, lsb, patchlists[plname]))
                except KeyError:
                    log.warning("Patch name list '%s' referenced by Patch bank '%s' not found.",
                                plname, pbname)
            else:
                patchlist = parse_patchnamelist(patchbank.find("./PatchNameList"))
                patchbanks.append(PatchBank(pbname, msb, lsb, patchlist))

        if patchbanks:
            namesets[nsname] = patchbanks

    return namesets


def write_xlsx(xlsxname, nameset):
    wb = Workbook()
    wb.add_named_style(colhead)

    # Use first work sheet named 'All' to collect all patches from all banks
    ws_all = wb.active
    ws_all.title = "All"
    ws_all_row = 1
    ws_all.freeze_panes = "A2"

    # Set column header styles for 'All' worksheet
    for col, (value, width, headstyle, _) in COLUMNS_ALL.items():
        ws_all[col + "1"] = value
        ws_all[col + "1"].style = headstyle
        ws_all.column_dimensions[col].width = width

    # Create one work sheet for each patch bank
    for pbnum, patchbank in enumerate(nameset):
        if not patchbank.patchlist:
            continue

        ws = wb.create_sheet(title=sanitize(patchbank.name))
        ws.freeze_panes = "A2"

        # Set column header styles for patch bank worksheet
        for col, (value, width, headstyle, _) in COLUMNS_BANK.items():
            ws[col + "1"] = value
            ws[col + "1"].style = headstyle
            ws.column_dimensions[col].width = width

        for row, patch in enumerate(patchbank.patchlist, start=2):
            if patchbank.name != "GM":
                ws_all.append((
                    patchbank.name,
                    patchbank.msb,
                    patchbank.lsb,
                    patch.program,
                    patch.category,
                    patch.name
                ))
                ws_all_row += 1

                # Set cell styles for row in 'All' work sheet
                for column in COLUMNS_ALL:
                    ws_all["%s%i" % (column, ws_all_row)].style = COLUMNS_ALL[column][3]

            ws.append((patch.program, patch.category, patch.name))

            # Set cell styles for row in patch bank work sheet
            for column in COLUMNS_BANK:
                ws["%s%i" % (column, row)].style = COLUMNS_BANK[column][3]

    # Set auto filter for 'All' work sheet
    ws_all.auto_filter.ref = ws_all.dimensions

    log.info("Writing '%s'...", xlsxname)
    wb.save(xlsxname)


def main(args=None):
    ap = argparse.ArgumentParser(prog="midnam2xlsx", description=__doc__.splitlines()[0])
    ap.add_argument("-v", "--verbose", action="store_true", help="Be verbose")
    ap.add_argument("midnam", type=open, help="MIDNAM input file")

    args = ap.parse_args(args)

    logging.basicConfig(
        format="%(name)s: %(levelname)s - %(message)s",
        level=logging.DEBUG if args.verbose else logging.INFO,
    )

    with args.midnam:
        namesets = parse_midnam(args.midnam)

    for nsnum, (nsname, nameset) in enumerate(namesets.items(), start=1):
        # Generate filename from nameset name
        basename = sanitize(nsname).strip()

        if basename:
            xlsxname = basename + ".xlsx"
        else:
            xlsxname = "nameset-%02i.xlsx" % nsmum

        write_xlsx(xlsxname, nameset)


if __name__ == "__main__":
    sys.exit(main() or 0)
