"""Openpyxl named cell style definitions."""

from openpyxl.styles import (
    NamedStyle,
    PatternFill,
    Border,
    Side,
    Alignment,
    Protection,
    Font,
)


align_center = Alignment(horizontal="center")
align_left = Alignment(horizontal="left")

bd = Side(style="thin", color="000000")
bd_all = Border(left=bd, top=bd, right=bd, bottom=bd)

fill_gray = PatternFill(fill_type="solid", start_color="dddddd", end_color="dddddd")
fill_white = PatternFill(fill_type="solid", start_color="ffffff", end_color="ffffff")

font_bold = Font(name="Liberation Sans", bold=True, size=11)
font_normal = Font(name="Liberation Sans", bold=False, size=11)

colhead = NamedStyle("colhead")
colhead.border = bd_all
colhead.fill = fill_gray
colhead.font = font_bold
colhead.alignment = align_center

numcell = NamedStyle("numcell")
numcell.alignment = align_center
numcell.border = bd_all
numcell.fill = fill_white
numcell.font = font_normal

txtcell = NamedStyle("txtcell")
txtcell.alignment = align_left
txtcell.border = bd_all
txtcell.fill = fill_white
txtcell.font = font_bold
